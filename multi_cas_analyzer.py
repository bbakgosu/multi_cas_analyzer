# Made by Chanju
import os
import sys
from collections import defaultdict
from Bio.SeqIO.QualityIO import FastqGeneralIterator
import regex
from Bio import Align
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

MINIMUM_FREQUENCY = 2



# def make_alignment(ref, read):
#     aligner = Align.PairwiseAligner()
#     aligner.mode = "global"
#     aligner.open_gap_score = -10
#     aligner.extend_gap_score = -0.5

#     alignments = aligner.align(ref, read)
#     alignment = alignments[0]
#     return alignment

def make_alignment(ref, read):
    aligner = Align.PairwiseAligner()
    aligner.mode = "global"
    aligner.open_gap_score = -10
    aligner.extend_gap_score = -0.5

    alignments = aligner.align(ref, read)
    alignment = alignments[0]
    return alignment

def make_alignment_string(ref, read):
    """Helper function to create alignment string for backward compatibility"""
    alignment = make_alignment(ref, read)
    
    # 시퀀스 번호 없이 정렬만 출력
    ref_aligned, read_aligned, matches = [], [], []

    ref_pos, read_pos = 0, 0
    for (ref_start, ref_end), (read_start, read_end) in zip(alignment.aligned[0], alignment.aligned[1]):
        # gap before the match
        while ref_pos < ref_start:
            ref_aligned.append(ref[ref_pos])
            read_aligned.append('-')
            matches.append(' ')
            ref_pos += 1
        while read_pos < read_start:
            ref_aligned.append('-')
            read_aligned.append(read[read_pos])
            matches.append(' ')
            read_pos += 1
        # match part
        for i in range(ref_start, ref_end):
            ref_aligned.append(ref[i])
            read_aligned.append(read[read_pos])
            if ref[i] == read[read_pos]:
                matches.append('|')
            else:
                matches.append(' ')
            ref_pos += 1
            read_pos += 1

    # 남은 부분 (gap at the end)
    while ref_pos < len(ref):
        ref_aligned.append(ref[ref_pos])
        read_aligned.append('-')
        matches.append(' ')
        ref_pos += 1
    while read_pos < len(read):
        ref_aligned.append('-')
        read_aligned.append(read[read_pos])
        matches.append(' ')
        read_pos += 1

    ref_str = ''.join(ref_aligned)
    match_str = ''.join(matches)
    read_str = ''.join(read_aligned)

    return f"{ref_str}\n{match_str}\n{read_str}"




def rc(seq):
    return seq.translate(str.maketrans("ATCGN", "TAGCN"))[::-1]



def get_indicator(seq, spacer, comparison_range=70):
    """
    Reference sequence로부터 indicator 시퀀스를 찾아냄.
    comparison_range보다 짧은 경우에도 최대한 근접하게 12bp 추출
    """
    spacer_pos = seq.find(spacer)
    if spacer_pos == -1:
        return None

    # Cas9 cleavage site 위치
    cleave_site = spacer_pos + len(spacer) - 3

    # forward indicator (upstream)
    fwd_start = max(cleave_site - comparison_range, 0)
    fwd_end = min(fwd_start + 12, len(seq))  # sequence 끝을 넘지 않도록
    fwd_indicator = seq[fwd_start:fwd_end]

    # reverse indicator (downstream)
    rev_end = min(cleave_site + comparison_range, len(seq))
    rev_start = max(rev_end - 12, 0)
    rev_indicator = seq[rev_start:rev_end]

    return fwd_indicator, rev_indicator




def crop_sequence(seq, fwd_indicator, rev_indicator, error=1):
    fwd_match = regex.search(f"({fwd_indicator}){{e<={error}}}", seq)
    rev_match = regex.search(f"({rev_indicator}){{e<={error}}}", seq)
    if (fwd_match and rev_match) and (fwd_match.start() < rev_match.start()):
        return seq[fwd_match.end():rev_match.start()]
    else:
        return None



def split_alignment(alignment, n=5):
    """split alignment into n parts in case of alignment is too long
    microsoft excel only display 1024 characters in a cell"""
    # alignment 객체를 문자열로 변환
    if isinstance(alignment, str):
        alignment_str = alignment
    else:
        alignment_str = str(alignment)
    alignment_list = alignment_str.split("\n")
    ref_seq, alignment_pattern, read_seq = alignment_list[0], alignment_list[1], alignment_list[2]
    seq_len = len(ref_seq)
    split_len = (seq_len // n) + 1
    split_alignment_list = []
    for i in range(n):
        split_alignment_list.append([
            ref_seq[i*split_len:(i+1)*split_len] + "\n" +
            alignment_pattern[i*split_len:(i+1)*split_len] + "\n" + 
            read_seq[i*split_len:(i+1)*split_len] + "\n"
        ])
    return split_alignment_list

def analyze_edit_pattern(alignment, wt_range):
    """
    Enhanced function to analyze detailed editing patterns
    Returns tuple: (classification, pattern_description, edit_size)
    """
    # Use make_alignment_string for consistent formatting
    alignment_str = make_alignment_string(alignment.sequences[0], alignment.sequences[1])
    alignment_lines = alignment_str.split("\n")
    target_aligned, align_diagram, query_aligned = alignment_lines[0], alignment_lines[1], alignment_lines[2]
    
    # Check for changes in WT range
    has_changes = False
    insertions = []
    deletions = []
    substitutions = []
    
    # Ensure wt_range is valid
    max_len = min(len(target_aligned), len(query_aligned), len(align_diagram))
    valid_range = [i for i in wt_range if 0 <= i < max_len]
    
    for i in valid_range:
        if align_diagram[i] != "|":
            has_changes = True
            if target_aligned[i] == "-":
                # Insertion in query
                insertions.append((i, query_aligned[i]))
            elif query_aligned[i] == "-":
                # Deletion in query
                deletions.append((i, target_aligned[i]))
            else:
                # Substitution
                substitutions.append((i, target_aligned[i], query_aligned[i]))
    
    if not has_changes:
        return "WT", "Wild Type", 0
    
    # Analyze patterns - prioritize simple edits over complex
    if insertions and not deletions and not substitutions:
        inserted_seq = ''.join([base for _, base in insertions])
        return "insertion", f"Insertion: {inserted_seq}", len(insertions)
    elif deletions and not insertions and not substitutions:
        deleted_seq = ''.join([base for _, base in deletions])
        return "deletion", f"Deletion: {deleted_seq}", len(deletions)
    elif substitutions and not insertions and not deletions:
        sub_patterns = [f"{ref}>{alt}" for _, ref, alt in substitutions]
        return "substitution", f"Substitution: {', '.join(sub_patterns)}", len(substitutions)
    else:
        # Check if it's primarily one type with minor other changes
        if len(insertions) > len(deletions) + len(substitutions):
            inserted_seq = ''.join([base for _, base in insertions])
            return "insertion", f"Insertion: {inserted_seq}", len(insertions)
        elif len(deletions) > len(insertions) + len(substitutions):
            deleted_seq = ''.join([base for _, base in deletions])
            return "deletion", f"Deletion: {deleted_seq}", len(deletions)
        else:
            # Complex edit (combination)
            pattern_parts = []
            if insertions:
                inserted_seq = ''.join([base for _, base in insertions])
                pattern_parts.append(f"Ins:{inserted_seq}")
            if deletions:
                deleted_seq = ''.join([base for _, base in deletions])
                pattern_parts.append(f"Del:{deleted_seq}")
            if substitutions:
                sub_patterns = [f"{ref}>{alt}" for _, ref, alt in substitutions]
                pattern_parts.append(f"Sub:{','.join(sub_patterns)}")
            
            total_size = len(insertions) + len(deletions) + len(substitutions)
            return "complex", f"Complex: {'; '.join(pattern_parts)}", total_size

def classify_read(alignment, wt_range):
    """
    Simple classification function for backward compatibility
    """
    # alignment가 문자열이면 그대로 사용, 객체이면 make_alignment_string 사용
    if isinstance(alignment, str):
        alignment_str = alignment
    else:
        alignment_str = make_alignment_string(alignment.sequences[0], alignment.sequences[1])
    
    alignment_lines = alignment_str.split("\n")
    target_aligned, align_diagram, query_aligned = alignment_lines[0], alignment_lines[1], alignment_lines[2]
    for i in wt_range:
        if i >= len(align_diagram):
            continue
        if align_diagram[i] != "|":
            if target_aligned[i] == "-":
                return "insertion"
            else:
                return "deletion"
    return "WT"
            

def get_edit_range_alignment(alignment, wt_range):
    """
    Extract only the edit_range portion of the alignment for display
    """
    # If alignment is already a string, use it directly
    if isinstance(alignment, str):
        alignment_str = alignment
    else:
        # Use make_alignment_string for consistent formatting
        alignment_str = make_alignment_string(alignment.sequences[0], alignment.sequences[1])
    
    alignment_lines = alignment_str.split("\n")
    target_aligned, align_diagram, query_aligned = alignment_lines[0], alignment_lines[1], alignment_lines[2]
    
    # Get the range boundaries
    start_pos = max(0, min(wt_range))
    end_pos = min(len(target_aligned), max(wt_range) + 1)
    
    # Extract the relevant portion
    target_part = target_aligned[start_pos:end_pos]
    align_part = align_diagram[start_pos:end_pos]
    query_part = query_aligned[start_pos:end_pos]
    
    return f"{target_part}\n{align_part}\n{query_part}"

def create_excel_report(pattern_data, output_path, target_name, wt_range):
    """
    Create Excel file with separate sheets for each edit type
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors for different edit types
    colors = {
        'WT': 'C8E6C9',
        'insertion': 'FFE0B2', 
        'deletion': 'FFCDD2',
        'substitution': 'E1BEE7',
        'complex': 'B3E5FC'
    }
    
    # Create sheets for each edit type
    for edit_type, patterns in pattern_data.items():
        if not patterns:  # Skip empty categories
            continue
            
        ws = wb.create_sheet(title=edit_type.capitalize())
        
        # Create headers - simplified to show alignment and count
        headers = ['Edit Range Alignment', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=colors.get(edit_type, 'FFFFFF'), 
                                  end_color=colors.get(edit_type, 'FFFFFF'), 
                                  fill_type='solid')
        
        # Set column widths
        ws.column_dimensions['A'].width = 50  # Alignment column
        ws.column_dimensions['B'].width = 15  # Count column
        
        # Sort patterns by count (descending)
        sorted_patterns = sorted(patterns.items(), key=lambda x: x[1]['count'], reverse=True)
        
        # Add data rows
        for row, (pattern, data) in enumerate(sorted_patterns, 2):
            # Get edit range alignment if sequence exists
            if 'alignment' in data and data['alignment']:
                edit_range_alignment = get_edit_range_alignment(data['alignment'], wt_range)
                cell = ws.cell(row=row, column=1, value=edit_range_alignment)
                cell.font = Font(name="Consolas", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Add count
            ws.cell(row=row, column=2, value=data['count'])
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title='Summary', index=0)
    summary_ws.cell(row=1, column=1, value='Target').font = Font(bold=True)
    summary_ws.cell(row=1, column=2, value=target_name)
    
    summary_ws.cell(row=3, column=1, value='Edit Type').font = Font(bold=True)
    summary_ws.cell(row=3, column=2, value='Total Count').font = Font(bold=True)
    summary_ws.cell(row=3, column=3, value='Unique Patterns').font = Font(bold=True)
    
    row = 4
    for edit_type, patterns in pattern_data.items():
        if patterns:
            total_count = sum(data['count'] for data in patterns.values())
            unique_patterns = len(patterns)
            
            summary_ws.cell(row=row, column=1, value=edit_type.capitalize())
            summary_ws.cell(row=row, column=2, value=total_count)
            summary_ws.cell(row=row, column=3, value=unique_patterns)
            
            # Color code the summary
            for col in range(1, 4):
                summary_ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=colors.get(edit_type, 'FFFFFF'),
                    end_color=colors.get(edit_type, 'FFFFFF'),
                    fill_type='solid'
                )
            row += 1
    
    # Auto-adjust summary column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    try:
        wb.save(output_path)
        print(f"Excel report saved to {output_path}")
    except PermissionError:
        # If file is open, try with a different name
        import time
        backup_path = output_path.replace('.xlsx', f'_backup_{int(time.time())}.xlsx')
        wb.save(backup_path)
        print(f"Excel report saved to {backup_path} (original file was in use)")

def main():
    input_folder = sys.argv[1]
    parameter_file = sys.argv[2]
    edit_range = int(sys.argv[3])


    fastq_files = [f for f in os.listdir(input_folder) if f.endswith(".fastq") or f.endswith(".fastqjoin")]
    parameter_dict = defaultdict(dict)
    os.makedirs(os.path.join(input_folder, "output"), exist_ok=True)
    with open(parameter_file) as f:
        lines = f.readlines()
        for i in range(0, len(lines), 3):
            key = i // 2
            parameter_dict[key]["name"] = lines[i].strip()
            parameter_dict[key]["ref"] = lines[i+1].strip().upper()
            parameter_dict[key]["spacer"] = lines[i + 2].strip().upper()

    # Create summary file for all files
    summary_output = os.path.join(input_folder, "output", "summary.txt")
    f_summary = open(summary_output, "w")
    f_summary.write("Filename\tTotal\tWT\tInsertion\tDeletion\tSubstitution\tComplex\n")

    for fastq_file in fastq_files:
        output_txt = os.path.join(input_folder, "output", fastq_file + ".txt")
        f_out = open(output_txt, "w")
        for key, value in parameter_dict.items():
            target_name = value["name"]
            ref_seq = value["ref"]
            spacer = value["spacer"]
            indicators = get_indicator(ref_seq, spacer)
            if indicators is None:
                indicators = get_indicator(ref_seq, rc(spacer))
                if indicators is None:
                    print(f"Error: No indicator found for {key} in {fastq_file}")
                    continue
            fwd_indicator, rev_indicator = indicators
            # print wtrange sequence
            ref_seq = crop_sequence(ref_seq, fwd_indicator, rev_indicator)
            if ref_seq.find(spacer) != -1:
                cut_site = ref_seq.find(spacer) + len(spacer) - 3
            else:
                cut_site = ref_seq.find(rc(spacer)) + 3
            # cut_site = ref_seq.find(spacer) + len(spacer) - 3
            wt_range = list(range(cut_site - edit_range, cut_site + edit_range))
            result_dict = {"With_both_indicator": 0, "WT": 0, "insertion": 0, "deletion": 0, "substitution": 0, "complex": 0}
            
            # Detailed pattern tracking for Excel export
            pattern_data = {
                "WT": {},
                "insertion": {},
                "deletion": {},
                "substitution": {},
                "complex": {}
            }
            
            with open(os.path.join(input_folder, fastq_file)) as f:
                sequence_dict = defaultdict(int)
                for title, seq, qual in FastqGeneralIterator(f):
                    cropped_seq = crop_sequence(seq, fwd_indicator, rev_indicator)
                    if cropped_seq:
                        sequence_dict[cropped_seq] += 1
            
            # Filter sequences with minimum frequency
            sequence_dict = {k: v for k, v in sequence_dict.items() if v >= MINIMUM_FREQUENCY}

            for seq, count in sequence_dict.items():
                alignment_result = make_alignment(ref_seq, seq)
                classification, pattern_description, edit_size = analyze_edit_pattern(alignment_result, wt_range)
                
                
                # Update simple counts
                result_dict[classification] += count
                result_dict["With_both_indicator"] += count
                
                # Track detailed patterns for Excel
                if pattern_description not in pattern_data[classification]:
                    # Store the alignment string for consistency
                    alignment_string = make_alignment_string(alignment_result.sequences[0], alignment_result.sequences[1])
                    pattern_data[classification][pattern_description] = {
                        'count': 0,
                        'description': pattern_description,
                        'edit_size': edit_size,
                        'sequence': seq,
                        'alignment': alignment_string  # alignment 문자열 저장
                    }
                pattern_data[classification][pattern_description]['count'] += count
            f_out.write(f"Target_name\t{target_name}\n")
            f_out.write("Total_reads\tWT\tInsertion\tDeletion\tSubstitution\tComplex\n")

            # Write counts
            f_out.write(f"{result_dict['With_both_indicator']}\t{result_dict['WT']}\t{result_dict['insertion']}\t{result_dict['deletion']}\t{result_dict['substitution']}\t{result_dict['complex']}\n")

            # Write percentages
            total = result_dict['With_both_indicator']
            if total > 0:
                wt_pct = result_dict['WT'] / total
                ins_pct = result_dict['insertion'] / total
                del_pct = result_dict['deletion'] / total
                sub_pct = result_dict['substitution'] / total
                cpx_pct = result_dict['complex'] / total
                f_out.write(f"1\t{wt_pct:.4f}\t{ins_pct:.4f}\t{del_pct:.4f}\t{sub_pct:.4f}\t{cpx_pct:.4f}\n")
            else:
                f_out.write("1\t0\t0\t0\t0\t0\n")
            
            # Generate Excel report for this target
            excel_output = os.path.join(input_folder, "output", f"{fastq_file}_{target_name}_patterns.xlsx")
            create_excel_report(pattern_data, excel_output, target_name, wt_range)

        # Write to summary file - counts
        total = result_dict['With_both_indicator']
        f_summary.write(f"{fastq_file}\t{total}\t{result_dict['WT']}\t{result_dict['insertion']}\t{result_dict['deletion']}\t{result_dict['substitution']}\t{result_dict['complex']}\n")

        # Write to summary file - percentages
        if total > 0:
            wt_pct = result_dict['WT'] / total
            ins_pct = result_dict['insertion'] / total
            del_pct = result_dict['deletion'] / total
            sub_pct = result_dict['substitution'] / total
            cpx_pct = result_dict['complex'] / total
            f_summary.write(f"{fastq_file}\t1\t{wt_pct:.4f}\t{ins_pct:.4f}\t{del_pct:.4f}\t{sub_pct:.4f}\t{cpx_pct:.4f}\n")
        else:
            f_summary.write(f"{fastq_file}\t1\t0\t0\t0\t0\t0\n")

        f_out.close()
        print(f"Results written to {output_txt}")
        print(f"\nSummary for {fastq_file}:")
        print(f"{'Total:':<15} {total:>10}")
        if total > 0:
            wt_pct = result_dict['WT'] / total * 100
            ins_pct = result_dict['insertion'] / total * 100
            del_pct = result_dict['deletion'] / total * 100
            sub_pct = result_dict['substitution'] / total * 100
            cpx_pct = result_dict['complex'] / total * 100
            print(f"{'WT:':<15} {result_dict['WT']:>10} ({wt_pct:>6.2f}%)")
            print(f"{'Insertion:':<15} {result_dict['insertion']:>10} ({ins_pct:>6.2f}%)")
            print(f"{'Deletion:':<15} {result_dict['deletion']:>10} ({del_pct:>6.2f}%)")
            print(f"{'Substitution:':<15} {result_dict['substitution']:>10} ({sub_pct:>6.2f}%)")
            print(f"{'Complex:':<15} {result_dict['complex']:>10} ({cpx_pct:>6.2f}%)")
        else:
            print(f"{'WT:':<15} {result_dict['WT']:>10} (  0.00%)")
            print(f"{'Insertion:':<15} {result_dict['insertion']:>10} (  0.00%)")
            print(f"{'Deletion:':<15} {result_dict['deletion']:>10} (  0.00%)")
            print(f"{'Substitution:':<15} {result_dict['substitution']:>10} (  0.00%)")
            print(f"{'Complex:':<15} {result_dict['complex']:>10} (  0.00%)")
        print()

    f_summary.close()
    print(f"\nSummary file written to {summary_output}")
        

                    
                

if __name__ == "__main__":
    main()                
